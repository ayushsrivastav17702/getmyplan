"""
Data Upload API routes — drop-in replacement with 75-error validation.
Uses get_db() for tenant-aware MongoDB access.
"""
from fastapi import APIRouter, UploadFile, File, Request, BackgroundTasks, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Optional
import uuid
import os
import io
import logging
from datetime import datetime, timedelta, timezone

from services.upload_service import UniversalUploadService, compute_file_hash
from multi_tenant.tenant_db import tenant_context
from services.cache_service import invalidate_for_upload, get_tenant_id as _cache_tid

logger = logging.getLogger(__name__)

router = APIRouter(tags=["upload"])

UPLOAD_DIR = "/tmp/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

MAX_FILE_SIZE_MB = 50
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

# Per-tenant upload lock (in-memory — sufficient for single-node)
import asyncio
_upload_locks: dict[str, asyncio.Lock] = {}


def _get_upload_lock(tenant_id: str) -> asyncio.Lock:
    if tenant_id not in _upload_locks:
        _upload_locks[tenant_id] = asyncio.Lock()
    return _upload_locks[tenant_id]


def _get_db():
    """Import here to avoid circular imports."""
    from server import get_db
    return get_db()


def _get_tenant_id():
    ctx = tenant_context.get()
    return ctx.tenant_id if ctx else "default"


def _get_user_email():
    """Best-effort user email extraction from tenant context."""
    return "system"


# ============================================================
# UPLOAD ENDPOINTS
# ============================================================

@router.post("/daily-sales")
async def upload_daily_sales(
    request: Request,
    file: UploadFile = File(...),
    replace_existing: bool = False,
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "daily_sales", replace_existing)


@router.post("/store-inventory")
async def upload_store_inventory(
    request: Request,
    file: UploadFile = File(...),
    replace_existing: bool = False,
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "store_inventory", replace_existing)


@router.post("/warehouse-inventory")
async def upload_warehouse_inventory(
    request: Request,
    file: UploadFile = File(...),
    replace_existing: bool = False,
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "warehouse_inventory", replace_existing)


@router.post("/sku-master")
async def upload_sku_master(
    request: Request,
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "sku_master", replace_existing=True)


@router.post("/store-master")
async def upload_store_master(
    request: Request,
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "store_master", replace_existing=True)


@router.post("/warehouse-master")
async def upload_warehouse_master(
    request: Request,
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "warehouse_master", replace_existing=True)


@router.post("/style-master")
async def upload_style_master(
    request: Request,
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "style_master", replace_existing=True)


@router.post("/cogs")
async def upload_cogs(
    request: Request,
    file: UploadFile = File(...),
    replace_existing: bool = False,
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "cogs", replace_existing)


@router.post("/planogram")
async def upload_planogram(
    request: Request,
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "planogram", replace_existing=True)


@router.post("/open-orders")
async def upload_open_orders(
    request: Request,
    file: UploadFile = File(...),
    replace_existing: bool = False,
    background_tasks: BackgroundTasks = None,
):
    return await _handle_upload(file, "open_orders", replace_existing)


# ============================================================
# VALIDATE-ONLY ENDPOINT
# ============================================================

@router.post("/{upload_type}/validate")
async def validate_file(
    upload_type: str,
    file: UploadFile = File(...),
):
    """Validate a file without saving to database."""
    normalized = upload_type.replace("-", "_")
    return await _handle_upload(file, normalized, replace_existing=False, validate_only=True)


# ============================================================
# STATUS & HISTORY ENDPOINTS
# ============================================================

@router.get("/history/days")
async def get_previous_days_history(days: int = 7):
    """Get upload status for previous days — per-day breakdown."""
    db = _get_db()
    tenant_id = _get_tenant_id()
    today = datetime.now(timezone.utc).date()
    result = []

    for i in range(1, days + 1):
        date = today - timedelta(days=i)
        date_str = date.strftime("%Y-%m-%d")

        uploads = {}
        for upload_type in ["daily_sales", "store_inventory", "warehouse_inventory", "cogs", "open_orders"]:
            doc = await db.upload_history.find_one(
                {"tenant_id": tenant_id, "upload_type": upload_type, "upload_date": date_str, "status": "completed"},
                {"_id": 0},
            )
            uploads[upload_type] = doc is not None

        if i == 1:
            label = "Yesterday"
        elif i < 7:
            label = date.strftime("%A")
        else:
            label = date.strftime("%b %d")

        if any(uploads.values()) or i <= 3:
            result.append({
                "date": date_str,
                "label": label,
                "uploads": uploads,
                "has_data": any(uploads.values()),
            })

    return {"days": result}


@router.get("/history")
async def get_upload_history(
    upload_type: Optional[str] = None,
    days: int = 7,
):
    """Get upload history grouped by date."""
    db = _get_db()
    tenant_id = _get_tenant_id()
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")

    query = {"tenant_id": tenant_id, "upload_date": {"$gte": start_date}}
    if upload_type:
        query["upload_type"] = upload_type

    history = await db.upload_history.find(query, {"_id": 0}).sort("upload_date", -1).to_list(200)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).strftime("%Y-%m-%d")

    grouped = {}
    for item in history:
        date_key = item.get("upload_date", "")
        if date_key == today:
            label = "Today"
        elif date_key == yesterday:
            label = "Yesterday"
        else:
            try:
                label = datetime.strptime(date_key, "%Y-%m-%d").strftime("%a, %b %d")
            except Exception:
                label = date_key

        if date_key not in grouped:
            grouped[date_key] = {"date": date_key, "label": label, "uploads": {}}

        uploaded_at = item.get("uploaded_at")
        time_str = uploaded_at.strftime("%I:%M %p") if hasattr(uploaded_at, "strftime") else str(uploaded_at)[:16]

        grouped[date_key]["uploads"][item["upload_type"]] = {
            "time": time_str,
            "rows": item.get("rows_uploaded", 0),
            "status": item.get("status", "completed"),
            "file_name": item.get("file_name", ""),
            "uploaded_by": item.get("uploaded_by", ""),
            "session_id": item.get("session_id", ""),
        }

    return {"history": list(grouped.values())}


@router.get("/daily-status")
async def get_daily_status():
    """Get today's upload status for the dashboard widget."""
    db = _get_db()
    tenant_id = _get_tenant_id()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    upload_types = ["daily_sales", "store_inventory", "warehouse_inventory", "cogs", "open_orders"]

    status = {}
    for ut in upload_types:
        upload = await db.upload_history.find_one(
            {"tenant_id": tenant_id, "upload_type": ut, "upload_date": today},
            {"_id": 0},
        )
        if upload:
            uploaded_at = upload.get("uploaded_at")
            time_str = uploaded_at.strftime("%I:%M %p") if hasattr(uploaded_at, "strftime") else ""
            status[ut] = {"uploaded": True, "time": time_str, "rows": upload.get("rows_uploaded", 0)}
        else:
            status[ut] = {"uploaded": False, "time": None, "rows": 0}

    return status


@router.get("/master-status")
async def get_master_status():
    """Get master data counts and last-updated timestamps."""
    db = _get_db()
    tenant_id = _get_tenant_id()
    result = {}

    COLLECTION_MAP = {
        "sku_master": "sku_ean_master",
        "store_master": "store_master",
        "warehouse_master": "warehouse_master",
        "style_master": "style_master",
        "planogram": "planogram",
    }

    for master_type in ["sku_master", "store_master", "warehouse_master", "style_master", "planogram"]:
        last = await db.upload_history.find_one(
            {"tenant_id": tenant_id, "upload_type": master_type, "status": "completed"},
            {"_id": 0},
            sort=[("uploaded_at", -1)],
        )
        coll_name = COLLECTION_MAP.get(master_type, master_type)
        # Try dedicated collection first — match tenant or no tenant_id (seeded data)
        count = await db[coll_name].count_documents(
            {"$or": [{"tenant_id": tenant_id}, {"tenant_id": {"$exists": False}}]}
        )
        if count == 0:
            doc = await db.uploaded_files.find_one({"file_type": master_type})
            if doc and "data" in doc:
                count = len(doc["data"])

        result[master_type] = {
            "count": count,
            "last_updated": last["uploaded_at"].strftime("%b %d") if last and hasattr(last.get("uploaded_at"), "strftime") else None,
        }

    return result


@router.get("/data-days")
async def get_data_days():
    """Return the number of distinct days of daily_sales data for the current tenant."""
    db = _get_db()
    tenant_id = _get_tenant_id()
    try:
        # Try with tenant_id filter first (V2 with tenant_id)
        days = await db.daily_sales.distinct("day", {"tenant_id": tenant_id})
        if not days:
            # Fallback: V2 without tenant_id (enterprise sample data)
            days = await db.daily_sales.distinct("day")
        if days:
            return {"days": len(days)}
        # Fallback to V1 uploaded_files
        doc = await db.uploaded_files.find_one({"file_type": "daily_sales"}, {"_id": 0, "data": 1})
        if doc and "data" in doc:
            day_set = set()
            for r in doc["data"]:
                d = r.get("day") or r.get("date")
                if d:
                    day_set.add(str(d))
            return {"days": len(day_set)}
        return {"days": 0}
    except Exception:
        return {"days": 0}


@router.get("/template/{upload_type}")
async def download_template(upload_type: str):
    """Download template pre-filled with tenant's actual SKUs and stores."""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    db = _get_db()
    _get_tenant_id()  # ensure tenant context
    wb = openpyxl.Workbook()
    ws = wb.active

    TEMPLATE_MAP = {
        "daily_sales": {
            "title": "Daily Sales",
            "headers": ["sku", "store_code", "day", "quantity", "revenue"],
        },
        "store_inventory": {
            "title": "Store Inventory",
            "headers": ["store_code", "sku", "closing_stock"],
        },
        "warehouse_inventory": {
            "title": "Warehouse Inventory",
            "headers": ["warehouse", "sku", "on_hand_qty", "available_qty"],
        },
        "sku_master": {
            "title": "SKU Master",
            "headers": ["sku", "product_name", "category"],
        },
        "store_master": {
            "title": "Store Master",
            "headers": ["store_code", "store_name"],
        },
        "warehouse_master": {
            "title": "Warehouse Master",
            "headers": ["warehouse", "warehouse_name", "online_fulfillment_flag"],
        },
        "style_master": {
            "title": "Style Master",
            "headers": ["style_code", "season", "category", "subcategory", "gender", "brand"],
        },
        "cogs": {
            "title": "COGS",
            "headers": ["transaction_date", "store_code", "sku_code", "cogs"],
        },
        "planogram": {
            "title": "Planogram",
            "headers": ["store_code", "category", "style_code", "norm_allocated", "repl_cycle_days", "cover_days", "top_seller_multiplier", "is_active"],
        },
        "open_orders": {
            "title": "Open Orders",
            "headers": ["order_date", "expected_delivery_date", "store_code", "sku_code", "order_quantity", "status", "source_type"],
        },
    }

    tmpl = TEMPLATE_MAP.get(upload_type)
    if not tmpl:
        raise HTTPException(400, f"Unknown upload type: {upload_type}")

    ws.title = tmpl["title"]
    for col, h in enumerate(tmpl["headers"], 1):
        ws.cell(row=1, column=col, value=h)

    # Add dropdowns from tenant master data
    if "sku" in tmpl["headers"]:
        skus = await db.uploaded_files.distinct("sku", {"file_type": "sku_master"})
        if not skus:
            skus = await db.uploaded_files.distinct("sku", {})
        if skus:
            sku_list = ",".join(str(s) for s in skus[:100])
            if len(sku_list) < 255:
                dv = DataValidation(type="list", formula1=f'"{sku_list}"')
                ws.add_data_validation(dv)
                col_idx = tmpl["headers"].index("sku") + 1
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                dv.add(f"{col_letter}2:{col_letter}10000")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={upload_type}_template.xlsx"},
    )


# ============================================================
# CORE HANDLER
# ============================================================

async def _handle_upload(file: UploadFile, upload_type: str, replace_existing: bool, validate_only: bool = False):
    """Internal handler — validates file, saves to DB if valid (unless validate_only)."""
    db = _get_db()
    tenant_id = _get_tenant_id()
    user_email = _get_user_email()
    session_id = f"UPL-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"

    lock = _get_upload_lock(tenant_id)
    if lock.locked() and not validate_only:
        return JSONResponse(content={
            "success": False,
            "errors": [{
                "code": "E057",
                "category": "file_structure",
                "message": "Another upload is in progress",
                "user_message": "Another upload is already being processed for this workspace. Please wait and try again.",
                "severity": "blocking",
            }],
            "total_rows": 0, "valid_rows": 0,
            "corrections": [], "warnings": [], "preview": [],
        })

    async with lock:
        return await _handle_upload_inner(file, upload_type, replace_existing, validate_only, db, tenant_id, user_email, session_id)


async def _handle_upload_inner(file, upload_type, replace_existing, validate_only, db, tenant_id, user_email, session_id):
    """Inner upload logic after acquiring the tenant lock."""

    # Plan limit check for store_master uploads
    if upload_type == "store_master" and not validate_only:
        from multi_tenant.tenant_db import get_shared_db
        from core.plan_access import check_plan_limit
        shared = get_shared_db()
        allowed, current, limit, plan = await check_plan_limit(shared, tenant_id, "stores")
        # We check after upload since replace_existing wipes old data
        # But we can warn if they're at the limit before uploading
        if not allowed and not replace_existing:
            return JSONResponse(content={
                "success": False,
                "errors": [{
                    "code": "E060",
                    "category": "plan_limit",
                    "message": "Store limit exceeded",
                    "user_message": f"Your {plan} plan allows {limit} stores (currently {current}). Upgrade to add more stores.",
                    "severity": "blocking",
                }],
                "total_rows": 0, "valid_rows": 0,
                "corrections": [], "warnings": [], "preview": [],
            })

    file_path = os.path.join(UPLOAD_DIR, f"{session_id}_{file.filename}")

    try:
        content = await file.read()

        # E049: File size check
        file_size_mb = len(content) / (1024 * 1024)
        if len(content) > MAX_FILE_SIZE_BYTES:
            return JSONResponse(content={
                "success": False,
                "errors": [{
                    "code": "E049",
                    "category": "file_structure",
                    "message": "File too large",
                    "user_message": f"Your file is {file_size_mb:.1f}MB. Maximum is {MAX_FILE_SIZE_MB}MB.",
                    "severity": "blocking",
                }],
                "total_rows": 0, "valid_rows": 0,
                "corrections": [], "warnings": [], "preview": [],
            })

        with open(file_path, "wb") as f:
            f.write(content)

        # E054: Duplicate file detection via hash
        file_hash = compute_file_hash(file_path)
        existing_upload = await db.upload_history.find_one(
            {"tenant_id": tenant_id, "file_hash": file_hash, "status": "completed"},
            {"_id": 0},
        )
        duplicate_warning = None
        if existing_upload:
            prev_date = existing_upload.get("uploaded_at")
            date_str = prev_date.strftime("%Y-%m-%d") if hasattr(prev_date, "strftime") else str(prev_date)[:10]
            time_str = prev_date.strftime("%I:%M %p") if hasattr(prev_date, "strftime") else ""
            duplicate_warning = {
                "code": "E054",
                "category": "duplicate",
                "message": "Same file uploaded twice",
                "user_message": f"This file was already uploaded on {date_str} at {time_str}.",
                "severity": "warning",
            }

        # Fetch master data for cross-validation (parallel)
        import asyncio as _aio
        master_skus, master_stores, master_warehouses = await _aio.gather(
            _get_master_skus(db, tenant_id),
            _get_master_stores(db, tenant_id),
            _get_master_warehouses(db, tenant_id),
        )

        service = UniversalUploadService(
            upload_type=upload_type,
            master_skus=master_skus,
            master_stores=master_stores,
            master_warehouses=master_warehouses,
            file_hash=file_hash,
        )

        result = service.process_file(file_path, file.filename)
        records = result.pop("data", None)

        # Add duplicate warning if found
        if duplicate_warning:
            result.setdefault("warnings", []).insert(0, duplicate_warning)

        result["validate_only"] = validate_only

        # If validation passed, save to database (skip if validate_only)
        if result["success"] and records and not validate_only:
            # Post-save store count check
            if upload_type == "store_master":
                from multi_tenant.tenant_db import get_shared_db
                from core.plan_access import check_plan_limit
                shared_db = get_shared_db()
                store_count = len(set(r.get("store_code") for r in records if r.get("store_code")))
                allowed, _, limit, plan = await check_plan_limit(shared_db, tenant_id, "stores")
                if store_count > limit:
                    result.setdefault("warnings", []).append({
                        "code": "W060",
                        "category": "plan_limit",
                        "message": f"Store limit: {store_count} stores uploaded but {plan} plan allows {limit}. Upgrade to keep all stores.",
                        "severity": "warning",
                    })

            saved = await _save_to_database(db, tenant_id, user_email, upload_type, records, replace_existing)
            result["saved"] = saved
            # Invalidate Redis caches for this tenant after successful upload
            try:
                invalidate_for_upload(tenant_id, upload_type)
                logger.info("Cache invalidated for tenant=%s upload_type=%s", tenant_id, upload_type)
            except Exception as cache_err:
                logger.warning("Cache invalidation failed (non-blocking): %s", cache_err)

        # Record history (skip for validate_only)
        if not validate_only:
            await db.upload_history.insert_one({
                "tenant_id": tenant_id,
                "upload_type": upload_type,
                "upload_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "file_name": file.filename,
                "file_hash": file_hash,
                "file_size_bytes": len(content),
                "rows_uploaded": result.get("total_rows", 0),
                "rows_valid": result.get("valid_rows", 0),
                "rows_with_warnings": len(result.get("warnings", [])),
                "rows_with_errors": len(result.get("errors", [])),
                "status": "completed" if result["success"] else "failed",
                "uploaded_by": user_email,
                "uploaded_at": datetime.now(timezone.utc),
                "session_id": session_id,
            })

        return JSONResponse(content=result)

    except Exception as e:
        logger.error(f"Upload handler error: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"success": False, "errors": [{"code": "FATAL", "message": str(e)}]},
        )
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


# ============================================================
# MASTER DATA FETCHERS
# ============================================================

async def _get_master_skus(db, tenant_id):
    """Get distinct SKUs from v2 sku_master collection OR v1 uploaded_files."""
    try:
        # Try v2 collection first
        skus = await db.sku_master.distinct("sku", {"tenant_id": tenant_id})
        if skus:
            return [str(s) for s in skus if s]

        # Fall back to v1 uploaded_files
        styles = await db.uploaded_files.find_one({"file_type": "sku_master"})
        if styles and "data" in styles:
            return list({str(r.get("sku", "")) for r in styles["data"] if r.get("sku")})
        styles = await db.uploaded_files.find_one({"file_type": "sku_ean_master"})
        if styles and "data" in styles:
            return list({str(r.get("ean", "")) for r in styles["data"] if r.get("ean")})
    except Exception:
        pass
    return []


async def _get_master_stores(db, tenant_id):
    """Get distinct stores from v2 store_master collection OR v1 uploaded_files."""
    try:
        # Try v2 collection first
        stores = await db.store_master.distinct("store_code", {"tenant_id": tenant_id})
        if stores:
            return [str(s) for s in stores if s]

        # Fall back to v1 uploaded_files
        doc = await db.uploaded_files.find_one({"file_type": "store_master"})
        if doc and "data" in doc:
            return list({str(r.get("store_code", "")) for r in doc["data"] if r.get("store_code")})
    except Exception:
        pass
    return []


async def _get_master_warehouses(db, tenant_id):
    """Get distinct warehouses from v2 warehouse_master collection OR v1 uploaded_files."""
    try:
        # Try v2 collection first
        warehouses = await db.warehouse_master.distinct("warehouse", {"tenant_id": tenant_id})
        if warehouses:
            return [str(w) for w in warehouses if w]

        # Fall back to v1 uploaded_files
        doc = await db.uploaded_files.find_one({"file_type": "warehouse_master"})
        if doc and "data" in doc:
            return list({str(r.get("warehouse", "")) for r in doc["data"] if r.get("warehouse")})
    except Exception:
        pass
    return []


async def _save_to_database(db, tenant_id, user_email, upload_type, records, replace_existing):
    """Save validated records to the appropriate collection."""
    collection_map = {
        "daily_sales": "daily_sales",
        "store_inventory": "store_inventory",
        "warehouse_inventory": "warehouse_inventory",
        "sku_master": "sku_master",
        "store_master": "store_master",
        "warehouse_master": "warehouse_master",
        "style_master": "style_master",
        "cogs": "cogs",
        "planogram": "planogram",
        "open_orders": "open_orders",
    }

    collection_name = collection_map.get(upload_type)
    if not collection_name:
        return False

    collection = db[collection_name]
    now_str = datetime.now(timezone.utc).isoformat()

    # Bulk enrich records (single pass)
    for record in records:
        record["tenant_id"] = tenant_id
        record["uploaded_at"] = now_str
        record["uploaded_by"] = user_email

    if replace_existing:
        # Batch delete by collecting unique keys first, then single bulk delete
        if upload_type == "daily_sales":
            days = {r.get("day") for r in records if r.get("day")}
            if days:
                await collection.delete_many({"tenant_id": tenant_id, "day": {"$in": list(days)}})
        elif upload_type == "cogs":
            dates = {r.get("transaction_date") for r in records if r.get("transaction_date")}
            if dates:
                await collection.delete_many({"tenant_id": tenant_id, "transaction_date": {"$in": list(dates)}})
        elif upload_type in ["open_orders", "store_inventory", "warehouse_inventory",
                             "sku_master", "store_master", "warehouse_master", "style_master", "planogram"]:
            await collection.delete_many({"tenant_id": tenant_id})

    if records:
        # Larger batches + unordered for max throughput
        BATCH = 5000
        for i in range(0, len(records), BATCH):
            await collection.insert_many(records[i:i + BATCH], ordered=False)

    return True



# ============================================================
# PREVIEW ENDPOINT
# ============================================================

@router.get("/preview/{upload_type}")
async def preview_data(upload_type: str, request: Request):
    """Return first 10 rows of an uploaded collection for preview."""
    db = _get_db()
    VALID = ["sku_master", "store_master", "warehouse_master", "style_master",
             "planogram", "daily_sales", "store_inventory", "warehouse_inventory",
             "cogs", "open_orders"]
    slug = upload_type.replace("-", "_")
    if slug not in VALID:
        raise HTTPException(400, f"Unknown upload type: {upload_type}")

    cursor = db[slug].find({}, {"_id": 0, "tenant_id": 0, "uploaded_by": 0, "uploaded_at": 0}).limit(10)
    preview = await cursor.to_list(10)
    total = await db[slug].estimated_document_count()

    # V1 fallback: check uploaded_files if V2 collection is empty
    if not preview:
        v1_doc = await db.uploaded_files.find_one({"file_type": slug}, {"_id": 0})
        if v1_doc and v1_doc.get("data"):
            preview = v1_doc["data"][:10]
            # Clean system fields from each row
            for row in preview:
                row.pop("_id", None)
                row.pop("tenant_id", None)
            total = len(v1_doc["data"])

    return {"preview": preview, "total": total, "type": slug}


# ============================================================
# SAMPLE DATA LOADER — Enterprise Scale (~220K rows)
# ============================================================

# In-memory job tracker (lightweight — only stores status dicts)
_seed_jobs = {}


@router.post("/load-sample-data")
async def load_sample_data(request: Request, background_tasks: BackgroundTasks):
    """Start sample data seeding in background. Returns immediately with job_id."""
    db = _get_db()

    # Check if force reseed requested
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    force = body.get("force", False)

    existing = await db.daily_sales.estimated_document_count()
    if existing > 50 and not force:
        return {"success": False, "message": "This tenant already has data. Use force=true to reseed, or upload your own data."}

    # If force reseed, clear existing sample collections first
    if existing > 0 and force:
        for coll in ["daily_sales", "store_master", "style_master", "sku_ean_master",
                      "warehouse_master", "warehouse_inventory", "planogram_norms",
                      "store_classes", "store_wedge_results", "style_mix_results"]:
            await db[coll].delete_many({})

    job_id = str(uuid.uuid4())[:12]
    _seed_jobs[job_id] = {"status": "pending", "progress": 0, "step": "Starting..."}

    background_tasks.add_task(_run_seed, db, job_id)

    return {"success": True, "message": "Sample data loading started", "job_id": job_id}


@router.get("/seed-status/{job_id}")
async def get_seed_status(job_id: str):
    """Poll progress of a seeding job."""
    return _seed_jobs.get(job_id, {"status": "not_found"})


async def _run_seed(db, job_id: str):
    """Background task — generates all sample data in batches."""
    import random
    import math
    import asyncio
    import gc
    random.seed(42)
    now = datetime.now(timezone.utc)
    j = _seed_jobs[job_id]

    try:
        j.update(status="processing", progress=5, step="Creating styles & SKUs...")

        # ── STORE MASTER: 30 stores ──
        STORES = [
            {"store_code":"DEL-01","store_name":"Delhi Connaught Place","city":"Delhi","region":"North","channel":"EBO","area_sqft":3200,"tier":"A"},
            {"store_code":"DEL-02","store_name":"Delhi Saket Mall","city":"Delhi","region":"North","channel":"EBO","area_sqft":2800,"tier":"A"},
            {"store_code":"DEL-03","store_name":"Delhi Lajpat Nagar","city":"Delhi","region":"North","channel":"MBO","area_sqft":1800,"tier":"B"},
            {"store_code":"GGN-01","store_name":"Gurgaon Cyber Hub","city":"Gurgaon","region":"North","channel":"EBO","area_sqft":2500,"tier":"A"},
            {"store_code":"NOI-01","store_name":"Noida Sector 18","city":"Noida","region":"North","channel":"MBO","area_sqft":2000,"tier":"B"},
            {"store_code":"LKO-01","store_name":"Lucknow Hazratganj","city":"Lucknow","region":"North","channel":"MBO","area_sqft":1500,"tier":"C"},
            {"store_code":"JAI-01","store_name":"Jaipur MI Road","city":"Jaipur","region":"North","channel":"MBO","area_sqft":1600,"tier":"C"},
            {"store_code":"CHD-01","store_name":"Chandigarh Sector 17","city":"Chandigarh","region":"North","channel":"EBO","area_sqft":2200,"tier":"B"},
            {"store_code":"BLR-01","store_name":"Bangalore Indiranagar","city":"Bangalore","region":"South","channel":"EBO","area_sqft":3000,"tier":"A"},
            {"store_code":"BLR-02","store_name":"Bangalore Koramangala","city":"Bangalore","region":"South","channel":"EBO","area_sqft":2600,"tier":"A"},
            {"store_code":"CHE-01","store_name":"Chennai T Nagar","city":"Chennai","region":"South","channel":"EBO","area_sqft":2400,"tier":"B"},
            {"store_code":"CHE-02","store_name":"Chennai Express Avenue","city":"Chennai","region":"South","channel":"EBO","area_sqft":2100,"tier":"B"},
            {"store_code":"HYD-01","store_name":"Hyderabad Jubilee Hills","city":"Hyderabad","region":"South","channel":"EBO","area_sqft":2800,"tier":"A"},
            {"store_code":"HYD-02","store_name":"Hyderabad Banjara Hills","city":"Hyderabad","region":"South","channel":"MBO","area_sqft":1900,"tier":"B"},
            {"store_code":"COK-01","store_name":"Kochi MG Road","city":"Kochi","region":"South","channel":"MBO","area_sqft":1400,"tier":"C"},
            {"store_code":"COI-01","store_name":"Coimbatore RS Puram","city":"Coimbatore","region":"South","channel":"MBO","area_sqft":1300,"tier":"C"},
            {"store_code":"MUM-01","store_name":"Mumbai Linking Road","city":"Mumbai","region":"West","channel":"EBO","area_sqft":3500,"tier":"A"},
            {"store_code":"MUM-02","store_name":"Mumbai Phoenix Mall","city":"Mumbai","region":"West","channel":"EBO","area_sqft":3200,"tier":"A"},
            {"store_code":"MUM-03","store_name":"Mumbai Andheri","city":"Mumbai","region":"West","channel":"MBO","area_sqft":2000,"tier":"B"},
            {"store_code":"PUN-01","store_name":"Pune Koregaon Park","city":"Pune","region":"West","channel":"EBO","area_sqft":2400,"tier":"B"},
            {"store_code":"PUN-02","store_name":"Pune FC Road","city":"Pune","region":"West","channel":"MBO","area_sqft":1700,"tier":"C"},
            {"store_code":"AMD-01","store_name":"Ahmedabad CG Road","city":"Ahmedabad","region":"West","channel":"MBO","area_sqft":1800,"tier":"B"},
            {"store_code":"SUR-01","store_name":"Surat Athwa Gate","city":"Surat","region":"West","channel":"MBO","area_sqft":1500,"tier":"C"},
            {"store_code":"KOL-01","store_name":"Kolkata Park Street","city":"Kolkata","region":"East","channel":"EBO","area_sqft":2600,"tier":"A"},
            {"store_code":"KOL-02","store_name":"Kolkata South City","city":"Kolkata","region":"East","channel":"MBO","area_sqft":1800,"tier":"B"},
            {"store_code":"BHU-01","store_name":"Bhubaneswar Patia","city":"Bhubaneswar","region":"East","channel":"MBO","area_sqft":1300,"tier":"C"},
            {"store_code":"PAT-01","store_name":"Patna Boring Road","city":"Patna","region":"East","channel":"MBO","area_sqft":1200,"tier":"C"},
            {"store_code":"GHY-01","store_name":"Guwahati GS Road","city":"Guwahati","region":"East","channel":"MBO","area_sqft":1100,"tier":"C"},
            {"store_code":"IND-01","store_name":"Indore MG Road","city":"Indore","region":"Central","channel":"MBO","area_sqft":1600,"tier":"C"},
            {"store_code":"BHO-01","store_name":"Bhopal MP Nagar","city":"Bhopal","region":"Central","channel":"MBO","area_sqft":1400,"tier":"C"},
        ]
        store_codes = [s["store_code"] for s in STORES]
        tier_map = {s["store_code"]: s["tier"] for s in STORES}
        TIER_MULT = {"A": 2.0, "B": 1.0, "C": 0.5}

        STYLES = [
            {"style_code":"STYLE-TS-001","style_name":"Classic Cotton T-Shirt","category":"Apparel","sub_category":"T-Shirts","brand":"Nike","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-TS-002","style_name":"Premium Pima T-Shirt","category":"Apparel","sub_category":"T-Shirts","brand":"Ralph Lauren","gender":"Men","season":"All Year"},
            {"style_code":"STYLE-HD-001","style_name":"Pullover Hoodie","category":"Apparel","sub_category":"Hoodies","brand":"Adidas","gender":"Unisex","season":"Fall/Winter"},
            {"style_code":"STYLE-HD-002","style_name":"Zip-Up Hoodie","category":"Apparel","sub_category":"Hoodies","brand":"Puma","gender":"Women","season":"Fall/Winter"},
            {"style_code":"STYLE-JG-001","style_name":"Slim Fit Joggers","category":"Apparel","sub_category":"Joggers","brand":"Under Armour","gender":"Men","season":"All Year"},
            {"style_code":"STYLE-JG-002","style_name":"Cargo Joggers","category":"Apparel","sub_category":"Joggers","brand":"Levi's","gender":"Men","season":"Fall/Winter"},
            {"style_code":"STYLE-PL-001","style_name":"Classic Polo","category":"Apparel","sub_category":"Polo Shirts","brand":"Lacoste","gender":"Men","season":"Spring/Summer"},
            {"style_code":"STYLE-PL-002","style_name":"Performance Polo","category":"Apparel","sub_category":"Polo Shirts","brand":"Nike","gender":"Men","season":"All Year"},
            {"style_code":"STYLE-SN-001","style_name":"Court Sneakers","category":"Footwear","sub_category":"Sneakers","brand":"Reebok","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-SN-002","style_name":"Running Shoes","category":"Footwear","sub_category":"Sneakers","brand":"Nike","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-SN-003","style_name":"Training Shoes","category":"Footwear","sub_category":"Sneakers","brand":"Adidas","gender":"Men","season":"All Year"},
            {"style_code":"STYLE-CP-001","style_name":"Baseball Cap","category":"Accessories","sub_category":"Caps","brand":"New Era","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-CP-002","style_name":"Snapback Cap","category":"Accessories","sub_category":"Caps","brand":"Mitchell & Ness","gender":"Men","season":"All Year"},
            {"style_code":"STYLE-SK-001","style_name":"Athletic Socks 3-Pack","category":"Accessories","sub_category":"Socks","brand":"Puma","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-SK-002","style_name":"Casual Socks 3-Pack","category":"Accessories","sub_category":"Socks","brand":"Happy Socks","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-BG-001","style_name":"Laptop Backpack","category":"Accessories","sub_category":"Bags","brand":"Herschel","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-BG-002","style_name":"Weekender Duffel","category":"Accessories","sub_category":"Bags","brand":"The North Face","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-BT-001","style_name":"Insulated Bottle 500ml","category":"Accessories","sub_category":"Bottles","brand":"Hydro Flask","gender":"Unisex","season":"All Year"},
            {"style_code":"STYLE-SW-001","style_name":"Crewneck Sweatshirt","category":"Apparel","sub_category":"Hoodies","brand":"Champion","gender":"Unisex","season":"Fall/Winter"},
            {"style_code":"STYLE-TK-001","style_name":"Performance Tank Top","category":"Apparel","sub_category":"T-Shirts","brand":"Nike","gender":"Men","season":"Spring/Summer"},
        ]
        WINTER_STYLES = {"STYLE-HD-001","STYLE-HD-002","STYLE-JG-002","STYLE-SW-001"}
        TOPSELLERS = {"STYLE-TS-001","STYLE-SN-001","STYLE-JG-001","STYLE-HD-001","STYLE-CP-001"}

        COLORS = {
            "STYLE-TS-001":["BLK","WHT","GRY","NAV"],"STYLE-TS-002":["WHT","BLU","PNK"],
            "STYLE-HD-001":["GRY","BLK","NAV"],"STYLE-HD-002":["BLK","PNK","WHT"],
            "STYLE-JG-001":["BLK","NAV","GRY"],"STYLE-JG-002":["OLV","BLK","KHK"],
            "STYLE-PL-001":["NAV","WHT"],"STYLE-PL-002":["BLK","GRY"],
            "STYLE-SN-001":["WHT","BLK","GRY"],"STYLE-SN-002":["BLK","RED","WHT"],
            "STYLE-SN-003":["BLK","WHT"],
            "STYLE-CP-001":["BLK","NAV","RED","WHT"],"STYLE-CP-002":["BLK","GRY"],
            "STYLE-SK-001":["WHT","BLK"],"STYLE-SK-002":["WHT","BLK","GRY"],
            "STYLE-BG-001":["BLK","NAV"],"STYLE-BG-002":["BLK"],
            "STYLE-BT-001":["BLK","WHT"],
            "STYLE-SW-001":["GRY","BLK","NAV"],"STYLE-TK-001":["BLK","WHT"],
        }
        SIZES_MAP = {"Apparel":["S","M","L","XL"],"Footwear":["7","8","9","10","11"],"Accessories":["ONE"]}
        SOCKS_SIZES = ["S/M","M/L"]
        MRP_MAP = {"T-Shirts":999,"Hoodies":2499,"Joggers":1999,"Polo Shirts":1499,"Sneakers":3499,"Caps":799,"Socks":499,"Bags":2999,"Bottles":699}

        sku_docs = []
        ean_to_info = {}
        for style_def in STYLES:
            sc, sub_cat, cat = style_def["style_code"], style_def["sub_category"], style_def["category"]
            colors = COLORS.get(sc, ["BLK"])
            sizes = SOCKS_SIZES if sub_cat == "Socks" else SIZES_MAP.get(cat, ["ONE"])
            mrp = MRP_MAP.get(sub_cat, 999)
            cogs_pct = random.uniform(0.40, 0.60)
            for clr in colors:
                for sz in sizes:
                    ean = f"{sc}-{clr}-{sz}"
                    cost = round(mrp * cogs_pct, 2)
                    sku_docs.append({"ean":ean,"style":sc,"color":clr,"size":sz,"mrp":mrp,"cost":cost,"barcode":ean,"category":cat,"sub_category":sub_cat})
                    ean_to_info[ean] = {"style":sc,"size":sz,"mrp":mrp,"cost":cost,"sub_cat":sub_cat}
        all_eans = list(ean_to_info.keys())

        WAREHOUSES = [
            {"warehouse_code":"WH-NCR","warehouse_name":"Delhi NCR DC","city":"Delhi","region":"North","capacity":60000,"online_fulfillment":True},
            {"warehouse_code":"WH-BLR","warehouse_name":"Bangalore DC","city":"Bangalore","region":"South","capacity":45000,"online_fulfillment":True},
            {"warehouse_code":"WH-MUM","warehouse_name":"Mumbai DC","city":"Mumbai","region":"West","capacity":50000,"online_fulfillment":True},
            {"warehouse_code":"WH-KOL","warehouse_name":"Kolkata DC","city":"Kolkata","region":"East","capacity":30000,"online_fulfillment":True},
        ]
        PLANOGRAMS = [
            {"store_code":"MUM-01","category":"T-Shirts","style":"STYLE-TS-001","norm_allocated":12,"replenish_cycle_days":7,"cover_days":14,"topseller_multiplier":1.5},
            {"store_code":"MUM-01","category":"Hoodies","style":"STYLE-HD-001","norm_allocated":8,"replenish_cycle_days":14,"cover_days":21,"topseller_multiplier":1.3},
            {"store_code":"DEL-01","category":"T-Shirts","style":"STYLE-TS-001","norm_allocated":10,"replenish_cycle_days":7,"cover_days":14,"topseller_multiplier":1.5},
            {"store_code":"BLR-01","category":"Sneakers","style":"STYLE-SN-001","norm_allocated":6,"replenish_cycle_days":14,"cover_days":21,"topseller_multiplier":1.2},
            {"store_code":"MUM-01","category":"Caps","style":"STYLE-CP-001","norm_allocated":15,"replenish_cycle_days":5,"cover_days":10,"topseller_multiplier":1.0},
            {"store_code":"DEL-01","category":"Joggers","style":"STYLE-JG-001","norm_allocated":8,"replenish_cycle_days":10,"cover_days":14,"topseller_multiplier":1.3},
            {"store_code":"BLR-01","category":"Bags","style":"STYLE-BG-001","norm_allocated":4,"replenish_cycle_days":21,"cover_days":30,"topseller_multiplier":1.0},
        ]

        # ═══ INSERT MASTER DATA ═══
        j.update(progress=10, step="Inserting master data...")
        for coll_name, docs in [("style_master",STYLES),("sku_ean_master",sku_docs),("store_master",STORES),("warehouse_master",WAREHOUSES),("planogram",PLANOGRAMS)]:
            await db[coll_name].delete_many({})
            if docs:
                await db[coll_name].insert_many(docs)
        await asyncio.sleep(0)

        # ═══ DAILY SALES + COGS (batched) ═══
        j.update(progress=15, step="Generating 90 days of sales...")
        BATCH = 3000
        total_sales = 0
        sales_batch, cogs_batch = [], []
        await db.daily_sales.delete_many({})
        await db.cogs.delete_many({})

        for day_offset in range(90):
            dt = now - timedelta(days=day_offset)
            day_str = dt.strftime("%Y-%m-%d")
            dow = dt.weekday()
            month = dt.month
            weekend_mult = 1.5 if dow >= 4 else 1.0
            winter_month = month in (12, 1, 2)

            for store in store_codes:
                store_mult = TIER_MULT[tier_map[store]]
                n_skus = max(10, int(len(all_eans) * random.uniform(0.30, 0.50)))
                day_skus = random.sample(all_eans, k=min(n_skus, len(all_eans)))
                for ean in day_skus:
                    info = ean_to_info[ean]
                    base_qty = random.randint(1, 6)
                    mult = store_mult * weekend_mult
                    if info["style"] in TOPSELLERS:
                        mult *= 3.0
                    if winter_month and info["style"] in WINTER_STYLES:
                        mult *= 1.8
                    qty = max(1, int(base_qty * mult))
                    discount = random.uniform(0.75, 1.0)
                    revenue = round(qty * info["mrp"] * discount, 2)
                    sales_batch.append({"day":day_str,"store_code":store,"sku":ean,"style":info["style"],"quantity":qty,"revenue":revenue,"mrp":info["mrp"]})
                    cogs_batch.append({"day":day_str,"store_code":store,"sku":ean,"style":info["style"],"quantity":qty,"cogs":round(qty*info["cost"],2),"revenue":revenue})
                    total_sales += 1

                    if len(sales_batch) >= BATCH:
                        await db.daily_sales.insert_many(sales_batch, ordered=False)
                        await db.cogs.insert_many(cogs_batch, ordered=False)
                        sales_batch.clear()
                        cogs_batch.clear()
                        await asyncio.sleep(0.1)

            # Update progress + GC every 10 days
            pct = 15 + int((day_offset / 90) * 60)
            j.update(progress=pct, step=f"Sales day {day_offset + 1}/90...")
            if day_offset % 10 == 0:
                gc.collect()

        if sales_batch:
            await db.daily_sales.insert_many(sales_batch)
        if cogs_batch:
            await db.cogs.insert_many(cogs_batch)

        # ═══ STORE INVENTORY ═══
        j.update(progress=78, step="Creating store inventory...")
        inv_docs = []
        today_str = now.strftime("%Y-%m-%d")
        for store in store_codes:
            for ean in all_eans:
                roll = random.random()
                qty = 0 if roll < 0.10 else random.randint(1, 5) if roll < 0.25 else random.randint(10, 35) if roll < 0.85 else random.randint(50, 120)
                inv_docs.append({"store_code":store,"sku":ean,"ean":ean,"style":ean_to_info[ean]["style"],"size":ean_to_info[ean]["size"],"quantity":qty,"closing_stock":qty,"day":today_str})
        await db.store_inventory.delete_many({})
        for i in range(0, len(inv_docs), BATCH):
            await db.store_inventory.insert_many(inv_docs[i:i + BATCH])
            await asyncio.sleep(0)

        # ═══ WAREHOUSE INVENTORY ═══
        j.update(progress=88, step="Creating warehouse inventory...")
        wh_inv = []
        for wh in WAREHOUSES:
            for ean in all_eans:
                wh_inv.append({"warehouse_code":wh["warehouse_code"],"warehouse":wh["warehouse_code"],"sku":ean,"style":ean_to_info[ean]["style"],"size":ean_to_info[ean]["size"],"quantity":random.randint(20,300),"day":today_str})
        await db.warehouse_inventory.delete_many({})
        await db.warehouse_inventory.insert_many(wh_inv)

        # ═══ OPEN ORDERS ═══
        j.update(progress=93, step="Creating open orders...")
        open_orders = []
        for i in range(8):
            ean = random.choice(all_eans)
            open_orders.append({"order_id":f"PO-2026-{1001+i}","order_type":"vendor_po","sku":ean,"style":ean_to_info[ean]["style"],"quantity":random.randint(50,500),"destination":random.choice(["WH-NCR","WH-BLR","WH-MUM","WH-KOL"]),"eta_date":(now+timedelta(days=random.randint(3,14))).strftime("%Y-%m-%d"),"status":"in_transit"})
        for i in range(7):
            ean = random.choice(all_eans)
            open_orders.append({"order_id":f"IST-2026-{2001+i}","order_type":"warehouse_transfer","sku":ean,"style":ean_to_info[ean]["style"],"quantity":random.randint(10,80),"source":random.choice(["WH-NCR","WH-BLR","WH-MUM","WH-KOL"]),"destination":random.choice(store_codes),"eta_date":(now+timedelta(days=random.randint(1,5))).strftime("%Y-%m-%d"),"status":"in_transit"})
        await db.open_orders.delete_many({})
        await db.open_orders.insert_many(open_orders)

        # ═══ CACHE INVALIDATION ═══
        j.update(progress=97, step="Finalizing...")
        try:
            from services.cache_service import invalidate_tenant
            invalidate_tenant(_cache_tid())
        except Exception:
            pass

        j.update(status="completed", progress=100, step="Complete!", summary={
            "styles":len(STYLES),"skus":len(sku_docs),"stores":len(STORES),
            "sales_records":total_sales,"days_of_history":90,
        })
        logger.info(f"Sample data seeded: {total_sales} sales rows")

    except Exception as e:
        logger.error(f"Sample data seed failed: {e}")
        j.update(status="failed", error=str(e))
